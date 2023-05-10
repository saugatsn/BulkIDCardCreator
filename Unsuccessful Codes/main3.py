import os
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.utils import get_column_letter



# Constants
TEMPLATE_PATH = "id_card_template.png"
PHOTOS_FOLDER = "photos"
BARCODE_FOLDER = "barcode"
EXCEL_FILE = "student_data.xlsx"
OUTPUT_FOLDER = "path/to/output/folder"
FONT_FOLDER = "output"
FONT_SIZE = 16
FONT_COLOR = (0, 0, 0)  # Black color
ROTATION_ANGLE = 90  # Degrees


def generate_id_cards():
    # Load template image
    template = Image.open(TEMPLATE_PATH)
    draw = ImageDraw.Draw(template)

    # Load Excel file
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    worksheet = workbook.active

    # Get the column letters for each field
    field_columns = {
        "Name": "A",
        "CRN": "B",
        "Level": "C",
        "Program": "D",
        "Cell": "E",
        "DOB": "F",
        "Blood Group": "G",
        "Citizenship": "H",
        "Email": "I",
        "ValidTill": "J"
    }

    # Iterate over rows in the Excel file
    for row in worksheet.iter_rows(min_row=2):
        # Retrieve data from each cell
        data = {field: list(row)[field_columns[field]] for field in field_columns}

        # Generate barcode path
        barcode_path = os.path.join(BARCODE_FOLDER, f"{data['Name']}.png")

        # Load student photo
        photo_path = os.path.join(PHOTOS_FOLDER, f"{data['Name']}.jpg")
        photo = Image.open(photo_path)

        # Resize student photo to desired dimensions
        photo_width, photo_height = 100, 100  # Adjust as needed
        photo = photo.resize((photo_width, photo_height), Image.ANTIALIAS)

        # Create circular mask for student photo
        mask = Image.new("L", (photo_width, photo_height), 0)
        draw_mask = ImageDraw.Draw(mask)
        draw_mask.ellipse((0, 0, photo_width, photo_height), fill=255)

        # Apply mask to student photo
        masked_photo = Image.new("RGBA", (photo_width, photo_height))
        masked_photo.paste(photo, (0, 0), mask=mask)

        # Insert data into the template image
        font_path = os.path.join(FONT_FOLDER, "font.ttf")  # Adjust font name and extension as needed
        font = ImageFont.truetype(font_path, FONT_SIZE)

        # Insert student photo
        photo_x, photo_y = 50, 50  # Adjust coordinates as needed
        template.paste(masked_photo, (photo_x, photo_y))

        # Insert barcode
        barcode_x, barcode_y = 200, 200  # Adjust coordinates as needed
        barcode_image = Image.open(barcode_path)
        template.paste(barcode_image, (barcode_x, barcode_y))

        # Insert text fields
        text_x, text_y = 300, 300  # Adjust coordinates as needed

        for field, value in data.items():
            # Calculate the position of the text field
            field_x = text_x
            field_y = text_y

            # Set font color, size, and type
            field_font = ImageFont.truetype(font_path, FONT_SIZE)
            field_color = FONT_COLOR

            # Adjust coordinates and other properties based on the field
            if field == 'ValidTill':
                field_value = str(value)
                field_font = ImageFont.truetype(font_path, FONT_SIZE - 2)  # Adjust font size
                field_x += 10  # Adjust x-coordinate
                field_y += 10  # Adjust y-coordinate
                draw.text((field_x, field_y), field_value, font=field_font, fill=field_color)
            elif field == 'Name':
                field_value = str(value)
                field_font = ImageFont.truetype(font_path, FONT_SIZE + 20)  # Adjust font size
                field_x += 10  # Adjust x-coordinate
                field_y += 10  # Adjust y-coordinate
                draw.text((field_x, field_y), field_value, font=field_font, fill=field_color)
            else:
                field_value = str(value)
                draw.text((field_x, field_y), field_value, font=field_font, fill=field_color)

            # Adjust the position for the next field
            text_y += 30  # Adjust the y-coordinate for the next field

        # Rotate the ValidTill field by 90 degrees anticlockwise
        valid_till_x, valid_till_y = text_x, text_y  # Adjust the coordinates for the ValidTill field
        valid_till_font = ImageFont.truetype(font_path, FONT_SIZE - 2)  # Adjust font size
        valid_till_color = FONT_COLOR
        valid_till_value = str(data['ValidTill'])
        rotated_valid_till = template.copy()
        rotated_valid_till = rotated_valid_till.rotate(ROTATION_ANGLE, expand=True)
        draw_valid_till = ImageDraw.Draw(rotated_valid_till)
        draw_valid_till.text((valid_till_x, valid_till_y), valid_till_value, font=valid_till_font, fill=valid_till_color)

        # Save the generated ID card as a JPG file
        output_file = os.path.join(OUTPUT_FOLDER, f"{data['Name']}.jpg")
        rotated_valid_till.save(output_file, "JPEG")

        # Close the opened images
        template.close()
        rotated_valid_till.close()

    # Close the workbook
    workbook.close()


# Run the function to generate ID cards
generate_id_cards()
