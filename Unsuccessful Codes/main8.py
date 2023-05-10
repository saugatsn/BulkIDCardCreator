from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import os

# Constants
TEMPLATE_PATH = "template.png"
PHOTOS_FOLDER = "photos"
BARCODES_FOLDER = "barcodes"
EXCEL_FILE = "student_data.xlsx"
OUTPUT_FOLDER = "Output"

# Define font properties
NAME_FONT = ImageFont.truetype("swiss.ttf", 35)
DETAILS_FONT = ImageFont.truetype("myriad.otf", 20)
VALID_TILL_FONT = ImageFont.truetype("arial.ttf", 38)

# Define colors
NAME_COLOR = (19, 4, 112)  # Black
DETAILS_COLOR = (19, 4, 112)  # Dark Gray
VALID_TILL_COLOR = (19, 4, 112)  # Red

# Define text positions
NAME_POSITION = (200, 360)
DETAILS_START_POSITION = (290, 422)
DETAILS_VERTICAL_GAP = 41
VALID_TILL_POSITION = (600, 380)

# Define photo positions and dimensions
PHOTO_WIDTH = 183
PHOTO_HEIGHT = 183

# Rotate ValidTill text
ROTATION_ANGLE = 90

# Load template image
template = Image.open("id_card_template.png")

# Load Excel data
wb = load_workbook("student_data.xlsx")
sheet = wb.active

# Create output folder if it doesn't exist
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Iterate over rows in Excel sheet (excluding header)
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Extract data from the row
    name, crn, level, program, cell, dob, blood_group, citizenship, email, valid_till = row
    if name is None:
        break
    name = name.upper()
    # Create a copy of the template image
    id_card = template.copy()

    # Create an ImageDraw object to draw on the ID card
    draw = ImageDraw.Draw(id_card)

    # Calculate photo position to center along the x-axis
    photo_position = (int((id_card.width - PHOTO_WIDTH) / 2), 153)

    # Insert student photo
    photo_path = os.path.join(PHOTOS_FOLDER, f"{name}.jpg")
    photo = Image.open(photo_path)
    photo = photo.resize((PHOTO_WIDTH, PHOTO_HEIGHT), Image.ANTIALIAS)
    mask = Image.new("L", photo.size, 0)
    mask_draw = ImageDraw.Draw(mask)
    mask_draw.ellipse((0, 0, PHOTO_WIDTH, PHOTO_HEIGHT), fill=255)
    id_card.paste(photo, photo_position, mask)

    # Calculate barcode position to center along the x-axis
    barcode_width = 200
    barcode_height = 100
    barcode_position = (int((id_card.width - barcode_width) / 2), 780)

    # Insert barcode
    barcode_path = os.path.join(BARCODES_FOLDER, f"{name}.jpg")
    barcode = Image.open(barcode_path)
    barcode = barcode.resize((barcode_width, barcode_height), Image.ANTIALIAS)
    id_card.paste(barcode, barcode_position)

    # Insert student details
    # Calculate the width of the name text
    name_width, _ = draw.textsize(name, font=NAME_FONT)

    # Calculate the x-coordinate to center the name
    name_x = int((id_card.width - name_width) / 2)
    name_position = (name_x, NAME_POSITION[1])

    # Insert student name
    draw.text(name_position, name, font=NAME_FONT, fill=NAME_COLOR)

    details_position = DETAILS_START_POSITION
    details = [f"{crn}",
               f"{level}",
               f"{program}",
               f"{cell}",
               f"{dob}",
               f"{blood_group}",
               f"{citizenship}",
               f"{email}"]
    for detail in details:
        draw.text(details_position, detail, font=DETAILS_FONT, fill=DETAILS_COLOR)
        details_position = (details_position[0], details_position[1] + DETAILS_VERTICAL_GAP)

    # Rotate and insert ValidTill
    valid_till_text = f"{valid_till}"
    rotated_valid_till = VALID_TILL_FONT.getsize(valid_till_text)
    rotated_valid_till = Image.new("RGBA", rotated_valid_till, (255, 255, 255, 0))
    rotated_valid_till_draw = ImageDraw.Draw(rotated_valid_till)
    rotated_valid_till_draw.text((0, 0), valid_till_text, font=VALID_TILL_FONT, fill=VALID_TILL_COLOR)
    rotated_valid_till = rotated_valid_till.rotate(ROTATION_ANGLE, expand=True)
    id_card.paste(rotated_valid_till, VALID_TILL_POSITION, rotated_valid_till)

    # Save ID card as JPEG
    output_path = os.path.join(OUTPUT_FOLDER, f"{name}.jpg")
    id_card.save(output_path, "JPEG")

    print(f"ID card generated for {name}.")

print("ID card generation completed.")

